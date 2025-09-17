#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
af_weekly_report.py
Зависимости: requests, pandas, openpyxl
ENV: AF_API_TOKEN=<appsFlyer v2 token>

Что делает:
- Читает apps из app_id.json (формат, как дал Саша).
- Выгружает из AppsFlyer daily_report/v5 CSV за прошлую и позапрошлую неделю.
- Считает installs ТОЛЬКО по paid (строки с ненулевым Campaign).
- w2w = (week - prev_week)/prev_week * 100, если prev_week>0.
- Команды: префиксы из Campaign до первого разделителя ([ _- |:/()[] ]) с суммой installs; выводим в отчёт в виде списка по убыванию.
- Пишет один файл af_weekly_report.xlsx:
  Лист "Отчёт" — колонки по ТЗ:
    1) Название приложения
    2) Тематика (пусто, для ручного ввода)
    3) Кол-во инсталлов за прошлую неделю
    4) Динамика w2w (%)
    5) Какие команды льют трафик
  Лист "Summary" — общий итог:
    6) Общее кол-во установок
    7) Динамика w2w
"""

import os
import re
import json
import io
from pathlib import Path
from datetime import date, datetime, timedelta
from typing import Dict, List, Tuple, Optional

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter

AF_BASE = "https://hq1.appsflyer.com"
TEAM_SPLIT = re.compile(r"[\s_\-\|:/\[\]\(\)]+")  # разделители для префикса команды

# ----------- даты -----------
def week_range(end_day: Optional[date] = None) -> Tuple[date, date, date, date]:
    """
    Возвращает (start_prev, end_prev, start_curr, end_curr)
    По умолчанию текущая неделя — последние 7 дней, заканчивающиеся вчера.
    """
    end_curr = (end_day or date.today()) - timedelta(days=1)
    start_curr = end_curr - timedelta(days=6)
    end_prev = start_curr - timedelta(days=1)
    start_prev = end_prev - timedelta(days=6)
    return start_prev, end_prev, start_curr, end_curr

# ----------- загрузка -----------
def fetch_daily_report_csv(app_id: str, d_from: date, d_to: date, token: str, timezone: str = "UTC") -> pd.DataFrame:
    """
    daily_report/v5 возвращает CSV. Берём базовые поля.
    """
    url = f"{AF_BASE}/api/agg-data/export/app/{app_id}/daily_report/v5"
    params = {
        "from": str(d_from),
        "to": str(d_to),
        "timezone": timezone,
        "retargeting": "false",  # только UA
    }
    headers = {"Authorization": f"Bearer {token}"}
    r = requests.get(url, headers=headers, params=params, timeout=180)
    r.raise_for_status()
    # CSV -> DataFrame
    buf = io.BytesIO(r.content)
    df = pd.read_csv(buf)
    return normalize_df(df)

def normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    colmap = {
        "Date": "date",
        "Media Source (pid)": "media_source",
        "Campaign (c)": "campaign",
        "Installs": "installs",
        "Total Cost": "total_cost",
    }
    df = df.rename(columns={k: v for k, v in colmap.items() if k in df.columns})
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.date
    for c in ["installs", "total_cost"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    # органику отсеиваем по отсутствию campaign (paid = есть campaign)
    if "campaign" in df.columns:
        df["campaign"] = df["campaign"].where(df["campaign"].notna(), None)
    # заполним отсутствующие числовые
    if "installs" not in df.columns:
        df["installs"] = 0
    return df

# ----------- расчёты -----------
def extract_team_prefix(campaign: Optional[str]) -> Optional[str]:
    if not isinstance(campaign, str) or not campaign.strip():
        return None
    return (TEAM_SPLIT.split(campaign.strip()) or [None])[0]

def installs_and_teams_paid(df: pd.DataFrame, top_n: Optional[int] = None) -> Tuple[int, List[str]]:
    """
    Считаем installs только по paid (строки с непустым campaign).
    Формируем список команд (префиксов) по убыванию installs.
    """
    if df.empty:
        return 0, []
    paid = df[df["campaign"].notna()].copy()
    total_installs = int(paid["installs"].sum()) if "installs" in paid.columns else 0
    if total_installs == 0:
        return 0, []
    # команды
    paid["team"] = paid["campaign"].apply(extract_team_prefix)
    teams = (
        paid.groupby("team", dropna=True)["installs"]
        .sum()
        .sort_values(ascending=False)
    )
    team_list = [t for t, v in teams.items() if t and v > 0]
    if top_n:
        team_list = team_list[:top_n]
    return total_installs, team_list

def wow(curr: int, prev: int) -> Optional[float]:
    if prev is None or prev == 0:
        return None
    return round(100.0 * (curr - prev) / prev, 2)

# ----------- Excel -----------
def write_excel(report_rows: List[Dict], totals_curr: int, totals_prev: int, out_path: Path):
    wb = Workbook()

    # Лист "Отчёт"
    ws = wb.active
    ws.title = "Отчёт"
    cols = [
        "Название приложения",
        "Тематика",
        "Кол-во инсталлов за прошлую неделю",
        "Динамика w2w (%)",
        "Какие команды льют трафик",
    ]
    ws.append(cols)
    for j, col in enumerate(cols, start=1):
        c = ws.cell(1, j)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for row in report_rows:
        ws.append([
            row["app_name"],
            "",  # Тематика — вручную
            row["installs_curr"],
            row["w2w_pct"],  # временно как число процентов → ниже форматируем
            ", ".join(row["teams"]),
        ])

    # Формат процента
    col_idx = {c: i + 1 for i, c in enumerate(cols)}
    pcol = col_idx["Динамика w2w (%)"]
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, pcol).value
        if v is None or v == "":
            ws.cell(r, pcol).value = None
            continue
        ws.cell(r, pcol).value = float(v) / 100.0
        ws.cell(r, pcol).number_format = numbers.FORMAT_PERCENTAGE_00

    # ширины
    widths = {
        "Название приложения": 30,
        "Тематика": 16,
        "Кол-во инсталлов за прошлую неделю": 26,
        "Динамика w2w (%)": 16,
        "Какие команды льют трафик": 48,
    }
    for name, w in widths.items():
        letter = get_column_letter(col_idx[name])
        ws.column_dimensions[letter].width = w
    ws.auto_filter.ref = ws.dimensions

    # Лист "Summary"
    ws2 = wb.create_sheet("Summary")
    total_wow = wow(totals_curr, totals_prev)
    ws2.append(["Общее кол-во установок (прошлая неделя)", totals_curr])
    ws2.append(["Общее кол-во установок (предыдущая неделя)", totals_prev if totals_prev else "—"])
    ws2.append(["Динамика w2w общая (%)", total_wow if total_wow is not None else "—"])
    # формат %
    if total_wow is not None:
        ws2.cell(3, 2).value = total_wow / 100.0
        ws2.cell(3, 2).number_format = numbers.FORMAT_PERCENTAGE_00

    wb.save(out_path)

# ----------- main -----------
def main():
    import argparse

    ap = argparse.ArgumentParser(description="AppsFlyer weekly report (по ТЗ).")
    ap.add_argument("--apps", default="app_id.json", help="Путь к app_id.json")
    ap.add_argument("--end", help="Конечная дата недели (YYYY-MM-DD), по умолчанию вчера")
    ap.add_argument("--tz", default="UTC", help="timezone для отчёта (UTC|preferred|...). По умолчанию UTC")
    ap.add_argument("--top-teams", type=int, default=10, help="Сколько команд выводить (по installs, по убыванию)")
    ap.add_argument("--out", default="af_weekly_report.xlsx", help="Выходной XLSX")
    args = ap.parse_args()

    token = os.getenv("AF_API_TOKEN")
    if not token:
        raise SystemExit("AF_API_TOKEN не задан")

    apps_obj = json.loads(Path(args.apps).read_text(encoding="utf-8"))
    apps: List[Dict] = apps_obj.get("apps", [])

    end_day = datetime.strptime(args.end, "%Y-%m-%d").date() if args.end else None
    p_s, p_e, c_s, c_e = week_range(end_day)

    report_rows = []
    total_curr = 0
    total_prev = 0

    for app in apps:
        app_id = app["id"]
        app_name = app.get("name") or app_id

        # текущая неделя
        try:
            df_curr = fetch_daily_report_csv(app_id, c_s, c_e, token, timezone=args.tz)
        except Exception as e:
            print(f"[ERR] {app_name} curr week: {e}")
            df_curr = pd.DataFrame()

        # предыдущая неделя
        try:
            df_prev = fetch_daily_report_csv(app_id, p_s, p_e, token, timezone=args.tz)
        except Exception as e:
            print(f"[ERR] {app_name} prev week: {e}")
            df_prev = pd.DataFrame()

        inst_curr, teams_curr = installs_and_teams_paid(df_curr, top_n=args.top_teams)
        inst_prev, _ = installs_and_teams_paid(df_prev, top_n=None)

        total_curr += inst_curr
        total_prev += inst_prev

        row = {
            "app_name": app_name,
            "installs_curr": inst_curr,
            "w2w_pct": wow(inst_curr, inst_prev),
            "teams": teams_curr,
        }
        report_rows.append(row)

    # сортировка по инсталлам по убыванию (чтобы важные сверху)
    report_rows.sort(key=lambda r: r["installs_curr"], reverse=True)

    write_excel(report_rows, total_curr, total_prev, Path(args.out))
    print(f"OK -> {args.out}")
    print(f"Период: {c_s}—{c_e} | Предыдущая: {p_s}—{p_e}")

if __name__ == "__main__":
    main()
