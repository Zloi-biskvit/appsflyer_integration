#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# report_form.py — Формирование отчёта по ТЗ + общий summary

import re
import json
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional, Tuple, List, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter

# === Конфиг ===
REPORT_DIR = Path("../report")            # где лежат daily_report*.csv
OUT_XLSX   = REPORT_DIR / "summary_readable.xlsx"
APPS_CACHE = REPORT_DIR / "apps_cache.json"  # {<app_id>: "<app_name>"}
THEMES_JSON = Path("./themes.json")      # {"<app_id>|<name>": "Тематика", ...}

# === Парсинг имён файлов ===
def parse_fname(p: Path) -> Tuple[Optional[datetime], Optional[datetime], Optional[str], Optional[str], Optional[str]]:
    """
    Поддержка:
    - daily_reportYYYY-MM-DD_YYYY-MM-DD_<appId>__<slug>.csv  (новый)
    - daily_reportYYYY-MM-DD_YYYY-MM-DD_<App Name>.csv       (старый)
    Возврат: (d1, d2, app_id, app_key, display_hint)
    """
    m = re.match(r"^daily_report(\d{4}-\d{2}-\d{2})_(\d{4}-\d{2}-\d{2})_([^_]+)__(.+)\.csv$", p.name)
    if m:
        d1 = datetime.strptime(m.group(1), "%Y-%m-%d").date()
        d2 = datetime.strptime(m.group(2), "%Y-%m-%d").date()
        app_id = m.group(3)
        slug = m.group(4).replace("_", " ")
        return d1, d2, app_id, app_id, slug

    m2 = re.match(r"^daily_report(\d{4}-\d{2}-\d{2})_(\d{4}-\d{2}-\d{2})_(.+)\.csv$", p.name)
    if m2:
        d1 = datetime.strptime(m2.group(1), "%Y-%m-%d").date()
        d2 = datetime.strptime(m2.group(2), "%Y-%m-%d").date()
        name = m2.group(3)
        return d1, d2, None, "name::" + name.strip().lower(), name

    return None, None, None, None, None

# === Нормализация CSV ===
COLMAP = {
    "Date": "date",
    "Media Source (pid)": "media_source",
    "Campaign (c)": "campaign",
    "Impressions": "impressions",
    "Clicks": "clicks",
    "CTR": "ctr",
    "Installs": "installs",
    "Conversion Rate": "cr",
    "Sessions": "sessions",
    "Loyal Users": "loyal_users",
    "Loyal Users/Installs": "loyal_ratio",
    "Total Cost": "total_cost",
    "Average eCPI": "ecpi",
}
NUMCOLS = ["impressions","clicks","ctr","installs","cr","sessions","loyal_users","loyal_ratio","total_cost","ecpi"]

def normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.rename(columns={k:v for k,v in COLMAP.items() if k in df.columns})
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.date
    for c in NUMCOLS:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    if "media_source" in df.columns:
        df["media_source"] = df["media_source"].fillna("Organic")
    if "campaign" in df.columns:
        df["campaign"] = df["campaign"].where(df["campaign"].notna(), None)
    return df

# === Метрики недели ===
TEAM_SPLIT = re.compile(r"[\s_\-\|:/\[\]\(\)]+")
def extract_team(name: Optional[str]) -> str:
    if not isinstance(name, str) or not name.strip():
        return "Unattributed"
    return (TEAM_SPLIT.split(name.strip()) or ["Unattributed"])[0]

def week_stats_and_teams(df: pd.DataFrame, top_n=10) -> Tuple[int, Optional[float], List[str]]:
    installs = int(df.get("installs", pd.Series(dtype=float)).sum())
    # команды
    teams = []
    if "campaign" in df.columns and "installs" in df.columns:
        t = df.assign(team=df["campaign"].apply(extract_team)) \
              .groupby("team")["installs"].sum().sort_values(ascending=False)
        teams = [k for k, v in t.items() if v > 0][:top_n]
    return installs, None, teams  # второй элемент — не используется здесь

# === Формирование отчёта ===
def choose_best_file(paths: List[Path]) -> Path:
    return sorted(paths, key=lambda p: (p.stat().st_size, p.stat().st_mtime), reverse=True)[0]

def load_maps() -> Tuple[Dict[str,str], Dict[str,str]]:
    apps_map = json.loads(APPS_CACHE.read_text(encoding="utf-8")) if APPS_CACHE.exists() else {}
    themes_map = json.loads(THEMES_JSON.read_text(encoding="utf-8")) if THEMES_JSON.exists() else {}
    # ключи тематик делаем case-insensitive для имён
    themes_norm = {}
    for k, v in themes_map.items():
        themes_norm[(k if k.startswith("id:") else k.strip().lower())] = v
    return apps_map, themes_norm

def resolve_app_name(app_id: Optional[str], display_hint: Optional[str], apps_map: Dict[str,str]) -> str:
    if app_id and app_id in apps_map:
        return apps_map[app_id]
    if display_hint:
        return display_hint
    return app_id or "UnknownApp"

def resolve_theme(app_id: Optional[str], app_name: str, themes_map: Dict[str,str]) -> str:
    if app_id and ("id:"+app_id) in themes_map:
        return themes_map["id:"+app_id]
    key = app_name.strip().lower()
    return themes_map.get(key, "Не задано")

def build_report_rows() -> Tuple[pd.DataFrame, Dict[str,int], Dict[str,int]]:
    apps_map, themes_map = load_maps()
    csvs = list(REPORT_DIR.glob("daily_report*.csv"))
    if not csvs:
        raise SystemExit("Нет файлов daily_report*.csv в ./report")

    # 1) на каждый (app_key, end_date) — лучший файл
    buckets: Dict[Tuple[str, datetime], List[Path]] = {}
    file_info: Dict[Path, Tuple[datetime, datetime, Optional[str], str, Optional[str]]] = {}
    for p in csvs:
        d1, d2, app_id, app_key, disp = parse_fname(p)
        if d2 is None or app_key is None:
            continue
        file_info[p] = (d1, d2, app_id, app_key, disp)
        buckets.setdefault((app_key, d2), []).append(p)
    chosen = {k: choose_best_file(v) for k, v in buckets.items()}

    # 2) словарь для быстрого доступа к installs по ключу/дате
    installs_index: Dict[Tuple[str, datetime], int] = {}
    for (app_key, end_d), p in chosen.items():
        df = normalize_df(pd.read_csv(p))
        inst, _, _ = week_stats_and_teams(df)
        installs_index[(app_key, end_d)] = inst

    # 3) для каждого app_key берём самую свежую неделю
    latest_per_app: Dict[str, Tuple[datetime, Path]] = {}
    for (app_key, end_d), p in chosen.items():
        if (app_key not in latest_per_app) or (end_d > latest_per_app[app_key][0]):
            latest_per_app[app_key] = (end_d, p)

    # 4) собираем строки формы
    rows = []
    totals_curr = 0
    totals_prev = 0
    for app_key, (end_d, p) in sorted(latest_per_app.items(), key=lambda x: x[0]):
        d1, d2, app_id, _, disp = file_info[p]
        df = normalize_df(pd.read_csv(p))
        installs, _, teams = week_stats_and_teams(df)

        prev_end = d2 - timedelta(days=7)
        prev_installs = installs_index.get((app_key, prev_end))
        w2w_pct = round(100.0 * (installs - prev_installs) / prev_installs, 2) if prev_installs else None

        app_name = resolve_app_name(app_id, disp, apps_map)
        theme = resolve_theme(app_id, app_name, themes_map)

        rows.append({
            "Название приложения": app_name,
            "Тематика": theme,
            "Кол-во инсталлов за прошлую неделю": installs,
            "Динамика w2w (%)": w2w_pct,
            "Команды (по префиксу кампании)": ", ".join(teams) if teams else "нет данных",
            "Период": f"{d1}—{d2}",
            "app_id": app_id or app_key,          # служебные
            "source_file": p.name                  # служебные
        })

        totals_curr += installs
        if prev_installs:
            totals_prev += prev_installs

    df_form = pd.DataFrame(rows)
    return df_form, {"curr": totals_curr}, {"prev": totals_prev}

def save_excel_form(df_form: pd.DataFrame, totals_curr: int, totals_prev: int, path: Path):
    wb = Workbook()

    # --- Лист "Отчёт" ---
    ws = wb.active
    ws.title = "Отчёт"

    visible_cols = ["Название приложения","Тематика","Кол-во инсталлов за прошлую неделю","Динамика w2w (%)","Команды (по префиксу кампании)","Период"]
    ws.append(visible_cols)
    for j, col in enumerate(visible_cols, start=1):
        c = ws.cell(1, j); c.font = Font(bold=True); c.alignment = Alignment(horizontal="center")

    # проценты -> доли для правильного % формата
    for _, r in df_form.iterrows():
        row = [r[c] for c in visible_cols]
        ws.append(row)

    # формат столбца w2w как %
    col_map = {c: i+1 for i, c in enumerate(visible_cols)}
    if "Динамика w2w (%)" in col_map:
        col = col_map["Динамика w2w (%)"]
        for r in range(2, ws.max_row+1):
            v = ws.cell(r, col).value
            if v is None or v == "":
                continue
            try:
                ws.cell(r, col).value = float(v)/100.0
                ws.cell(r, col).number_format = numbers.FORMAT_PERCENTAGE_00
            except Exception:
                pass

    # ширины
    widths = {
        "Название приложения": 28,
        "Тематика": 18,
        "Кол-во инсталлов за прошлую неделю": 24,
        "Динамика w2w (%)": 16,
        "Команды (по префиксу кампании)": 44,
        "Период": 18
    }
    for c, w in widths.items():
        letter = get_column_letter(col_map[c])
        ws.column_dimensions[letter].width = w
    ws.auto_filter.ref = ws.dimensions

    # --- Лист "Summary" ---
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Всего инсталлов (текущая неделя)", totals_curr])
    w2w_total = round(100.0 * (totals_curr - totals_prev) / totals_prev, 2) if totals_prev else None
    ws2.append(["Всего инсталлов (пред. неделя)", totals_prev if totals_prev else "—"])
    ws2.append(["Динамика w2w общая (%)", w2w_total if w2w_total is not None else "—"])

    # формат %
    if w2w_total is not None:
        ws2.cell(3, 2).value = w2w_total/100.0
        ws2.cell(3, 2).number_format = numbers.FORMAT_PERCENTAGE_00

    # чисто служебный лист со ссылками на файлы (удобно для дебага)
    ws3 = wb.create_sheet("Tech")
    ws3.append(["app_id","Название приложения","Период","source_file"])
    for _, r in df_form.iterrows():
        ws3.append([r["app_id"], r["Название приложения"], r["Период"], r["source_file"]])

    wb.save(path)

def main():
    df_form, total_curr_dict, total_prev_dict = build_report_rows()
    save_excel_form(df_form, total_curr_dict["curr"], total_prev_dict["prev"], OUT_XLSX)
    print(f"OK -> {OUT_XLSX}")

if __name__ == "__main__":
    main()
